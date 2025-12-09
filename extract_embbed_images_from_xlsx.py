import xml.etree.ElementTree as ET
import zipfile
import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from pathlib import Path
import threading
import queue

# openpyxl导入将在create_excel_worksheet方法中处理
OPENPYXL_AVAILABLE = None  # 将在运行时确定

def win_path(path_str):
    """将路径字符串转换为Windows格式（如果适用）"""
    if os.path.sep == "\\":
        return f"{str(path_str).replace("/", '\\')}"
    return str(path_str)

class XLSXImageExtractorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("XLSX单元格内嵌入图像提取工具")
        self.root.geometry("800x600")
        
        # 设置样式
        self.root.configure(bg='#f0f0f0')
        
        # 创建主框架
        main_frame = ttk.Frame(root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 配置网格权重
        root.columnconfigure(0, weight=1)
        root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # 文件选择部分
        ttk.Label(main_frame, text="XLSX文件:", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        
        self.file_path_var = tk.StringVar()
        self.file_entry = ttk.Entry(main_frame, textvariable=self.file_path_var, width=50)
        self.file_entry.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.browse_file_btn = ttk.Button(main_frame, text="浏览...", command=self.browse_file)
        self.browse_file_btn.grid(row=1, column=2, padx=(5, 0), pady=(0, 10))
        
        # 输出目录选择部分
        ttk.Label(main_frame, text="输出目录:", font=('Arial', 10, 'bold')).grid(row=2, column=0, sticky=tk.W, pady=(0, 5))
        
        self.output_dir_var = tk.StringVar()
        self.output_entry = ttk.Entry(main_frame, textvariable=self.output_dir_var, width=50)
        self.output_entry.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.browse_output_btn = ttk.Button(main_frame, text="浏览...", command=self.browse_output_dir)
        self.browse_output_btn.grid(row=3, column=2, padx=(5, 0), pady=(0, 10))
        
        # 选项部分
        options_frame = ttk.LabelFrame(main_frame, text="选项", padding="10")
        options_frame.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.create_csv_var = tk.BooleanVar(value=False)
        self.create_csv_check = ttk.Checkbutton(options_frame, text="提取结果输出为CSV文件 (extracted_images.csv)", 
                                                variable=self.create_csv_var)
        self.create_csv_check.grid(row=0, column=0, sticky=tk.W)
        
        self.create_excel_var = tk.BooleanVar(value=True)
        self.create_excel_check = ttk.Checkbutton(options_frame, text="提取结果输出为Excel工作表 (extracted_images.xlsx)", 
                                                  variable=self.create_excel_var)
        self.create_excel_check.grid(row=1, column=0, sticky=tk.W, pady=(5, 0))
        
        self.extract_images_var = tk.BooleanVar(value=True)
        self.extract_images_check = ttk.Checkbutton(options_frame, text="提取图像文件到media目录", 
                                                    variable=self.extract_images_var)
        self.extract_images_check.grid(row=2, column=0, sticky=tk.W, pady=(5, 0))
        
        # 控制按钮
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=5, column=0, columnspan=3, pady=(10, 10))
        
        self.extract_btn = ttk.Button(button_frame, text="开始提取", command=self.start_extraction, width=15)
        self.extract_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        self.clear_btn = ttk.Button(button_frame, text="清空日志", command=self.clear_log, width=15)
        self.clear_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        self.exit_btn = ttk.Button(button_frame, text="退出", command=root.quit, width=15)
        self.exit_btn.pack(side=tk.LEFT)
        
        # 进度条
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(main_frame, variable=self.progress_var, maximum=100, length=400)
        self.progress_bar.grid(row=6, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # 状态标签
        self.status_var = tk.StringVar(value="就绪")
        self.status_label = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        self.status_label.grid(row=7, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # 日志区域
        ttk.Label(main_frame, text="日志输出:", font=('Arial', 10, 'bold')).grid(row=8, column=0, sticky=tk.W, pady=(0, 5))
        
        self.log_text = scrolledtext.ScrolledText(main_frame, width=70, height=15, wrap=tk.WORD)
        self.log_text.grid(row=9, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 配置网格权重使日志区域可扩展
        main_frame.rowconfigure(9, weight=1)
        
        # 消息队列用于线程安全更新UI
        self.message_queue = queue.Queue()
        
        # 定期检查消息队列
        self.check_queue()
        
        # 设置默认输出目录为桌面
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        self.output_dir_var.set(desktop_path)
    
    def browse_file(self):
        """浏览并选择XLSX文件"""
        file_path = filedialog.askopenfilename(
            title="选择XLSX文件",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if file_path:
            self.file_path_var.set(win_path(file_path))
            # 自动设置输出目录为文件所在目录
            if not self.output_dir_var.get():
                output_dir = os.path.abspath(os.path.dirname(file_path))
                self.output_dir_var.set(output_dir)
    
    def browse_output_dir(self):
        """浏览并选择输出目录"""
        dir_path = win_path(filedialog.askdirectory(title="选择输出目录"))
        if dir_path:
            self.output_dir_var.set(dir_path)
    
    def log_message(self, message):
        """向日志区域添加消息"""
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
    
    def clear_log(self):
        """清空日志区域"""
        self.log_text.delete(1.0, tk.END)
    
    def update_status(self, message):
        """更新状态标签"""
        self.status_var.set(message)
        self.root.update_idletasks()
    
    def update_progress(self, value):
        """更新进度条"""
        self.progress_var.set(value)
        self.root.update_idletasks()
    
    def check_queue(self):
        """定期检查消息队列并更新UI"""
        try:
            while True:
                msg_type, *args = self.message_queue.get_nowait()
                if msg_type == "log":
                    self.log_message(*args)
                elif msg_type == "status":
                    self.update_status(*args)
                elif msg_type == "progress":
                    self.update_progress(*args)
                elif msg_type == "enable_buttons":
                    self.enable_buttons(*args)
                elif msg_type == "show_message":
                    messagebox.showinfo(*args)
        except queue.Empty:
            pass
        finally:
            self.root.after(100, self.check_queue)
    
    def queue_message(self, msg_type, *args):
        """将消息放入队列"""
        self.message_queue.put((msg_type, *args))
    
    def enable_buttons(self, enable=True):
        """启用或禁用按钮"""
        state = tk.NORMAL if enable else tk.DISABLED
        self.extract_btn.config(state=state)
        self.browse_file_btn.config(state=state)
        self.browse_output_btn.config(state=state)
        self.clear_btn.config(state=state)
    
    def extract_cellimages_from_xlsx(self, xlsx_path):
        """
        从XML文件中提取xdr:cNvpr元素的name属性和a:blip元素的r:embed属性
        返回元组列表 [(ID1, image1), (ID2, image2), ...]
        """
        # 定义XML命名空间
        namespaces = {
            'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
            'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
            'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
            'etc': 'http://www.wps.cn/officeDocument/2017/etCustomData'
        }
        
        try:
            with zipfile.ZipFile(xlsx_path, 'r') as z:
                if 'xl/cellimages.xml' in z.namelist():
                    with z.open('xl/cellimages.xml') as f:
                        tree = ET.parse(f)
                        root = tree.getroot()
                        
                        result = []
                        
                        # 查找所有etc:cellImage元素
                        for cell_image in root.findall('etc:cellImage', namespaces):
                            # 查找xdr:cNvPr元素并获取name属性
                            c_nv_pr = cell_image.find('.//xdr:cNvPr', namespaces)
                            name = c_nv_pr.get('name') if c_nv_pr is not None else None
                            
                            # 查找a:blip元素并获取r:embed属性
                            blip = cell_image.find('.//a:blip', namespaces)
                            embed = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed') if blip is not None else None
                            
                            # get name and embed, remove 'rId' prefix from embed
                            # get image name from rId
                            if name is not None and embed is not None:
                                rid = embed.removeprefix('rId')
                                for fn in z.namelist():
                                    if fn.startswith(f"xl/media/image{rid}."):
                                        image_name = fn.removeprefix("xl/media/") 
                                        result.append((name, image_name))
                                        break
                        
                        return result
            
        except Exception as e:
            self.queue_message("log", f"错误: {e}")
            return []
    
    def extract_subdir_from_zip(self, zip_path: str, subdir: str, dest_dir: str):
        """
        Extract only the files inside 'subdir/' (or deeper) from the zip.
        
        Parameters:
            zip_path  – path to the .zip file
            subdir    – folder name inside the zip (e.g. "myfolder" or "path/to/myfolder")
            dest_dir  – where to extract on disk
        """
        # Make sure subdir ends with '/' so we match is exact
        if not subdir.endswith('/'):
            subdir += '/'
        
        os.makedirs(dest_dir, exist_ok=True)
        
        with zipfile.ZipFile(zip_path, 'r') as zf:
            files = [m for m in zf.namelist() if m.startswith(subdir)]
            total_files = len(files)
            
            for idx, member in enumerate(files):
                # Only extract files that are inside the desired subdirectory
                if member.startswith(subdir):
                    # Remove the leading subdir part so we don't create extra nesting
                    target_path = os.path.join(dest_dir, member[len(subdir):])
                    
                    # If it's a directory entry (ends with '/'), create the dir
                    if member.endswith('/'):
                        os.makedirs(target_path, exist_ok=True)
                    else:
                        # Ensure parent directory exists
                        os.makedirs(os.path.dirname(target_path), exist_ok=True)
                        # Extract the file
                        with zf.open(member) as source, open(target_path, "wb") as target:
                            target.write(source.read())
                    
                    # 更新进度
                    progress = (idx + 1) / total_files * 100
                    self.queue_message("progress", progress)
                    self.queue_message("log", f"提取文件: {member[len(subdir):]}")
    
    def create_csv(self, tuples_list, output_dir):
        """创建CSV文件来存储提取的图像信息"""
        try:
            csv_path = os.path.abspath(os.path.join(output_dir, 'extracted_images.csv'))
            
            with open(csv_path, 'w') as csv:
                # 写入CSV表头
                csv.write('ID,img_path\n')
                
                # 遍历所有图像信息并写入CSV
                for i, (ID, image) in enumerate(tuples_list):
                    # 构建图像文件的完整路径
                    img_path = os.path.abspath(os.path.join(output_dir, "media", image))
                    
                    # 处理Windows路径分隔符（将单个反斜杠替换为双反斜杠）
                    if os.path.sep == "\\":
                        img_path = f"{str(img_path).replace(os.sep, '\\\\')}"
                    
                    # 写入CSV行（使用双引号包裹字段值）
                    csv.write(f'"{ID}","{img_path}"\n')
            
            self.queue_message("log", f"CSV文件已保存到 {csv_path}")
            return True
            
        except Exception as e:
            self.queue_message("log", f"创建CSV文件时出错: {e}")
            return False
    
    def create_excel_worksheet(self, tuples_list, output_dir, media_dir):
        """创建Excel工作表来存储提取的图像信息"""
        try:
            # 尝试导入openpyxl
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side
            
            excel_path = os.path.abspath(os.path.join(output_dir, 'extracted_images.xlsx'))
            
            # 创建工作簿和工作表
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "提取的图像"
            
            # 设置标题行
            headers = ["序号", "图像名称", "图像文件", "完整路径", "提取时间"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 添加数据行
            from datetime import datetime
            extract_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            for row_idx, (image_name, image_file) in enumerate(tuples_list, 2):
                img_path = os.path.join(media_dir, image_file) if media_dir else image_file
                full_path = os.path.abspath(img_path) if os.path.exists(img_path) else f"未找到: {img_path}"
                
                ws.cell(row=row_idx, column=1, value=row_idx-1)  # 序号
                ws.cell(row=row_idx, column=2, value=image_name)  # 图像名称
                ws.cell(row=row_idx, column=3, value=image_file)  # 图像文件
                ws.cell(row=row_idx, column=4, value=full_path.replace("\\","\\\\"))   # 完整路径
                ws.cell(row=row_idx, column=5, value=extract_time) # 提取时间
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 添加边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=1, max_row=len(tuples_list)+1, max_col=5):
                for cell in row:
                    cell.border = thin_border
            
            # 保存工作簿
            wb.save(excel_path)
            self.queue_message("log", f"Excel工作表保存到 {excel_path}")
            return True
            
        except ImportError:
            self.queue_message("log", "警告: openpyxl库未安装，无法创建Excel文件")
            self.queue_message("log", "请使用以下命令安装: pip install openpyxl")
            return False
        except Exception as e:
            self.queue_message("log", f"创建Excel文件时出错: {e}")
            return False
    
    def extraction_thread(self):
        """执行提取操作的线程函数"""
        try:
            # 获取输入参数
            xlsx_path = self.file_path_var.get()
            output_dir = self.output_dir_var.get()
            create_csv = self.create_csv_var.get()
            create_excel = self.create_excel_var.get()
            extract_images = self.extract_images_var.get()
            
            # 验证输入
            if not xlsx_path or not os.path.exists(xlsx_path):
                self.queue_message("show_message", "错误", "请选择有效的XLSX文件")
                self.queue_message("enable_buttons", True)
                return
            
            if not output_dir:
                self.queue_message("show_message", "错误", "请选择输出目录")
                self.queue_message("enable_buttons", True)
                return
            
            # 创建输出目录
            os.makedirs(output_dir, exist_ok=True)
            
            self.queue_message("status", "正在提取图像信息...")
            self.queue_message("progress", 10)
            
            # 提取图像信息
            tuples_list = self.extract_cellimages_from_xlsx(xlsx_path)
            
            if not tuples_list:
                self.queue_message("log", "警告: 未找到任何图像")
                self.queue_message("status", "完成 - 未找到图像")
                self.queue_message("progress", 100)
                self.queue_message("enable_buttons", True)
                return
            
            self.queue_message("log", "提取的图像列表:")
            self.queue_message("log", "[")
            for i, (ID, rID) in enumerate(tuples_list):
                self.queue_message("log", f'"{ID}","{rID}"')
            self.queue_message("log", "]")
            
            self.queue_message("log", f"\n总共提取了 {len(tuples_list)} 个图像")
            self.queue_message("progress", 30)
            
            # 创建CSV文件
            if create_csv:
                self.queue_message("status", "正在创建CSV文件...")
                if self.create_csv(tuples_list, output_dir):
                    self.queue_message("progress", 50)
            
            # 创建Excel工作表
            if create_excel:
                self.queue_message("status", "正在创建Excel工作表...")
                media_dir =os.path.abspath(os.path.join(output_dir, 'media')) if extract_images else None
                if self.create_excel_worksheet(tuples_list, output_dir, media_dir):
                    self.queue_message("log", "Excel工作表创建成功")
                self.queue_message("progress", 70)
            
            # 提取图像文件
            if extract_images:
                self.queue_message("status", "正在提取图像文件...")
                media_dir = os.path.join(output_dir, 'media')
                self.extract_subdir_from_zip(xlsx_path, 'xl/media', media_dir)
                self.queue_message("log", f"图像文件已提取到 {media_dir}")
                self.queue_message("progress", 90)
            
            # 完成
            self.queue_message("status", f"完成 - 提取了 {len(tuples_list)} 个图像")
            self.queue_message("progress", 100)
            self.queue_message("log", "\n操作完成！")
            self.queue_message("show_message", "完成", f"成功提取了 {len(tuples_list)} 个图像")
            
        except Exception as e:
            self.queue_message("log", f"错误: {e}")
            self.queue_message("status", "错误")
            self.queue_message("show_message", "错误", f"提取过程中发生错误: {str(e)}")
        finally:
            self.queue_message("enable_buttons", True)
    
    def start_extraction(self):
        """开始提取过程"""
        # 禁用按钮防止重复点击
        self.enable_buttons(False)
        
        # 清空日志
        self.clear_log()
        
        # 重置进度条
        self.progress_var.set(0)
        
        # 在后台线程中执行提取操作
        thread = threading.Thread(target=self.extraction_thread, daemon=True)
        thread.start()

def main():
    root = tk.Tk()
    _app = XLSXImageExtractorGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()
